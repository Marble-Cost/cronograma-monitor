import io
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Colores
NAVY   = "003A70"
TEAL   = "00B5B0"
WHITE  = "FFFFFF"
LIGHT  = "F0F9FF"
GRAY   = "F1F5F9"
DKGRAY = "64748B"
GREEN  = "16A34A"
LGREEN = "DCFCE7"
YELLOW = "D97706"
LYELLOW= "FEF9C3"
LRED   = "FEE2E2"
LBLUE  = "CBD5E1"

def _fill(color):
    return PatternFill("solid", start_color=color, end_color=color)

def _font(bold=False, color=None, size=11, italic=False):
    return Font(bold=bold, color=color or "000000", size=size, italic=italic,
                name="Calibri")

def _border(thin=True):
    s = Side(style="thin" if thin else "medium", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _header_row(ws, row, cols_data, fill_color=NAVY, font_color=WHITE, height=22):
    ws.row_dimensions[row].height = height
    for col_idx, (col_letter, value, width) in enumerate(cols_data, 1):
        c = ws.cell(row=row, column=col_idx, value=value)
        c.fill    = _fill(fill_color)
        c.font    = _font(bold=True, color=font_color, size=10)
        c.alignment = _center(wrap=True)
        c.border  = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

# ═══════════════════════════════════════════════════════
# HOJA 1: Panel de Control
# ═══════════════════════════════════════════════════════
def _build_panel(wb, kpis, phase_progress, config, scenario, sheet_name_crono):
    ws = wb.create_sheet("Panel de Control", 0)
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("A1:S1")
    c = ws["A1"]
    c.value     = "COMPLIANCE MONITOR — SCHEDULE"
    c.fill      = _fill(NAVY)
    c.font      = _font(bold=True, color=WHITE, size=18)
    c.alignment = _center()
    ws.row_dimensions[1].height = 40

    # Subtítulo
    ws.merge_cells("A2:S2")
    c = ws["A2"]
    c.value     = f"Informe de Progreso · Escenario: {scenario} · Exportado: {date.today().strftime('%d/%m/%Y')}"
    c.fill      = _fill(TEAL)
    c.font      = _font(bold=False, color=WHITE, size=11)
    c.alignment = _center()
    ws.row_dimensions[2].height = 22

    # Separador
    ws.row_dimensions[3].height = 10

    # KPIs — fórmulas que apuntan a la hoja de cronograma
    ws.merge_cells("A4:C4")
    ws["A4"].value = "📊 RESUMEN EJECUTIVO"
    ws["A4"].font  = _font(bold=True, color=NAVY, size=12)
    ws.row_dimensions[4].height = 24

    kpi_data = [
        ("A5:D8",  "COMPLETADAS",  f"=COUNTIF('{sheet_name_crono}'!E:E,\"COMPLETADO\")", LGREEN, GREEN),
        ("E5:H8",  "EN PROGRESO",  f"=COUNTIF('{sheet_name_crono}'!E:E,\"EN PROGRESO\")", LYELLOW, YELLOW),
        ("I5:L8",  "PENDIENTES",   f"=COUNTIF('{sheet_name_crono}'!E:E,\"PENDIENTE\")", LBLUE, NAVY),
        ("M5:P8",  "% AVANCE",    f"=IFERROR(COUNTIF('{sheet_name_crono}'!E:E,\"COMPLETADO\")/25,0)", LIGHT, TEAL),
    ]

    for merge_range, label, formula, bg, fg in kpi_data:
        ws.merge_cells(merge_range)
        start_cell = merge_range.split(":")[0]
        c = ws[start_cell]
        c.fill      = _fill(bg)
        c.font      = _font(bold=True, color=fg, size=28)
        c.alignment = _center()
        c.border    = _border()
        if "%" in label:
            c.number_format = "0%"
        c.value = formula

        # Label debajo
        label_row = int(merge_range.split(":")[1][1:]) + 1
        label_col = merge_range.split(":")[0][0]
        end_col   = merge_range.split(":")[1][0]
        label_merge = f"{label_col}{label_row}:{end_col}{label_row}"
        try:
            ws.merge_cells(label_merge)
        except Exception:
            pass
        lc = ws[f"{label_col}{label_row}"]
        lc.value     = label
        lc.fill      = _fill(bg)
        lc.font      = _font(bold=True, color=fg, size=9)
        lc.alignment = _center()
        lc.border    = _border()

    ws.row_dimensions[5].height = 42
    ws.row_dimensions[6].height = 6
    ws.row_dimensions[7].height = 6
    ws.row_dimensions[8].height = 6
    ws.row_dimensions[9].height = 18

    # Tabla de avance por fase
    ws.row_dimensions[10].height = 10
    ws.merge_cells("A11:S11")
    ws["A11"].value     = "AVANCE POR FASE"
    ws["A11"].font      = _font(bold=True, color=NAVY, size=12)
    ws["A11"].fill      = _fill(GRAY)
    ws["A11"].alignment = _center()
    ws.row_dimensions[11].height = 22

    fase_headers = [("A", "FASE", 18), ("B", "NOMBRE", 36),
                    ("C", "COMPLETADAS", 15), ("D", "EN PROGRESO", 15),
                    ("E", "PENDIENTES", 13), ("F", "% COMPLETADO", 14)]
    _header_row(ws, 12, [(l, v, w) for l, v, w in fase_headers], NAVY, WHITE, 20)

    fases_info = [
        (0, "FASE 0 · Alineación y Arranque", 5),
        (1, "FASE 1 · Base de Datos", 6),
        (2, "FASE 2 · Backend Python", 5),
        (3, "FASE 3 · Frontend y Seguridad", 4),
        (4, "FASE 4 · Testing y Despliegue", 5),
    ]
    crono = sheet_name_crono

    for i, (fn, fname, total) in enumerate(fases_info):
        r = 13 + i
        ws.row_dimensions[r].height = 20
        bg = WHITE if i % 2 == 0 else GRAY

        vals = [
            f"FASE {fn}", fname,
            f"=COUNTIFS('{crono}'!B:B,\"FASE {fn}*\",'{crono}'!E:E,\"COMPLETADO\")",
            f"=COUNTIFS('{crono}'!B:B,\"FASE {fn}*\",'{crono}'!E:E,\"EN PROGRESO\")",
            f"=COUNTIFS('{crono}'!B:B,\"FASE {fn}*\",'{crono}'!E:E,\"PENDIENTE\")",
            f"=IFERROR(COUNTIFS('{crono}'!B:B,\"FASE {fn}*\",'{crono}'!E:E,\"COMPLETADO\")/{total},0)",
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill      = _fill(bg)
            c.font      = _font(size=10)
            c.alignment = _center() if ci != 2 else _left(wrap=False)
            c.border    = _border()
            if ci == 6:
                c.number_format = "0%"

    # Fila total
    r_total = 18
    ws.row_dimensions[r_total].height = 22
    for ci, val in enumerate([
        "TOTAL", f"Escenario: {scenario}",
        f"=COUNTIF('{crono}'!E:E,\"COMPLETADO\")",
        f"=COUNTIF('{crono}'!E:E,\"EN PROGRESO\")",
        f"=COUNTIF('{crono}'!E:E,\"PENDIENTE\")",
        f"=IFERROR(COUNTIF('{crono}'!E:E,\"COMPLETADO\")/25,0)",
    ], 1):
        c = ws.cell(row=r_total, column=ci, value=val)
        c.fill      = _fill(TEAL)
        c.font      = _font(bold=True, color=WHITE, size=10)
        c.alignment = _center() if ci != 2 else _left(wrap=False)
        c.border    = _border()
        if ci == 6:
            c.number_format = "0%"

    # Nota al pie
    ws.row_dimensions[20].height = 14
    ws.merge_cells("A20:S20")
    c = ws["A20"]
    c.value     = f"⚠ Los KPIs se calculan automáticamente. Si editas los estados en la hoja '{crono}', este panel se actualiza al recalcular (F9)."
    c.font      = _font(italic=True, color=DKGRAY, size=9)
    c.alignment = _left(wrap=False)

    # Anchos de columnas del panel
    for col in range(1, 20):
        ws.column_dimensions[get_column_letter(col)].width = 9

# ═══════════════════════════════════════════════════════
# HOJA 2: Cronograma
# ═══════════════════════════════════════════════════════
def _build_cronograma(wb, activities, config, scenario):
    sheet_name = f"Cronograma {scenario}"
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C4"

    # Título
    ws.merge_cells("A1:T1")
    c = ws["A1"]
    c.value     = f"CRONOGRAMA DE DESPLIEGUE — {scenario.upper()}"
    c.fill      = _fill(NAVY)
    c.font      = _font(bold=True, color=WHITE, size=14)
    c.alignment = _center()
    ws.row_dimensions[1].height = 32

    # Fecha inicio
    start = config.start_date
    ws.merge_cells("A2:T2")
    c = ws["A2"]
    inicio_txt = f"Fecha de inicio: {start.strftime('%d/%m/%Y')}" if start else "Fecha de inicio: No definida"
    c.value     = f"{inicio_txt}  ·  Exportado: {date.today().strftime('%d/%m/%Y')}  ·  Total: 25 actividades"
    c.fill      = _fill(TEAL)
    c.font      = _font(color=WHITE, size=10)
    c.alignment = _center()
    ws.row_dimensions[2].height = 18

    # Encabezados
    headers = [
        ("#",         4),  # A
        ("FASE",      22), # B
        ("ACTIVIDAD", 42), # C
        ("RESPONSABLE",14),# D
        ("ESTADO",    16), # E
        ("S.INI",     6),  # F
        ("S.FIN",     6),  # G
        ("S1",5),("S2",5),("S3",5),("S4",5),("S5",5),("S6",5),
        ("S7",5),("S8",5),("S9",5),("S10",5),("S11",5),("S12",5),
        ("ÚLTIMA OBSERVACIÓN", 35), # T
    ]
    ws.row_dimensions[3].height = 20
    for ci, (hdr, width) in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=hdr)
        c.fill      = _fill(NAVY)
        c.font      = _font(bold=True, color=WHITE, size=9)
        c.alignment = _center(wrap=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Data validations
    dv_status = DataValidation(
        type="list",
        formula1='"PENDIENTE,EN PROGRESO,COMPLETADO"',
        allow_blank=False, showErrorMessage=True,
        errorTitle="Estado inválido",
        error="Elige: PENDIENTE, EN PROGRESO o COMPLETADO"
    )
    dv_resp = DataValidation(
        type="list",
        formula1='"Desarrollador,TI,Ambos,Liderazgo"',
        allow_blank=False
    )
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_resp)

    # Datos
    sorted_acts = sorted(activities, key=lambda a: (a.fase_number, a.activity_number))
    STATUS_BG = {"PENDIENTE": LBLUE, "EN PROGRESO": LYELLOW, "COMPLETADO": LGREEN}

    obs_map = {}  # pre-loaded from activities — observations passed separately

    for i, act in enumerate(sorted_acts):
        r   = 4 + i
        bg  = "FFFFFF" if i % 2 == 0 else "F8FAFC"
        sbg = STATUS_BG.get(act.status, LBLUE)
        ws.row_dimensions[r].height = 18

        row_data = [
            (1,  act.activity_number, _center(), bg),
            (2,  act.fase_name,       _left(wrap=False), bg),
            (3,  act.activity_name,   _left(wrap=True),  bg),
            (4,  act.responsable,     _center(), bg),
            (5,  act.status,          _center(), sbg),
            (6,  act.week_start,      _center(), bg),
            (7,  act.week_end,        _center(), bg),
        ]
        for ci, val, aln, bg_col in row_data:
            c = ws.cell(row=r, column=ci, value=val)
            c.fill      = _fill(bg_col)
            c.font      = _font(size=9, bold=(ci==5))
            c.alignment = aln
            c.border    = _border()

        # Gantt columns (H=8 to S=19 → weeks 1-12)
        for w in range(1, 13):
            ci = 7 + w  # col 8 = S1, col 19 = S12
            in_range = act.week_start <= w <= act.week_end
            if in_range:
                gantt_bg = {"COMPLETADO": "4ADE80", "EN PROGRESO": "FCD34D", "PENDIENTE": "CBD5E1"}.get(act.status, "CBD5E1")
            else:
                gantt_bg = "FFFFFF" if i % 2 == 0 else "F8FAFC"
            c = ws.cell(row=r, column=ci, value="")
            c.fill      = _fill(gantt_bg)
            c.border    = _border()

        # Observación
        c = ws.cell(row=r, column=20, value=act.notes or "")
        c.fill      = _fill(bg)
        c.font      = _font(size=9, italic=True, color=DKGRAY)
        c.alignment = _left(wrap=True)
        c.border    = _border()

        # Dropdowns
        dv_status.add(ws.cell(row=r, column=5))
        dv_resp.add(ws.cell(row=r, column=4))

    # Conditional formatting para el Gantt — se actualiza si el usuario cambia el estado
    data_end = 3 + len(sorted_acts)
    for w in range(1, 13):
        col_letter = get_column_letter(7 + w)
        cf_range   = f"{col_letter}4:{col_letter}{data_end}"
        ref_row    = 4

        ws.conditional_formatting.add(cf_range, FormulaRule(
            formula=[f'AND($E{ref_row}="COMPLETADO",$F{ref_row}<={w},$G{ref_row}>={w})'],
            fill=_fill("4ADE80"), stopIfTrue=True
        ))
        ws.conditional_formatting.add(cf_range, FormulaRule(
            formula=[f'AND($E{ref_row}="EN PROGRESO",$F{ref_row}<={w},$G{ref_row}>={w})'],
            fill=_fill("FCD34D"), stopIfTrue=True
        ))
        ws.conditional_formatting.add(cf_range, FormulaRule(
            formula=[f'AND($E{ref_row}="PENDIENTE",$F{ref_row}<={w},$G{ref_row}>={w})'],
            fill=_fill("CBD5E1"), stopIfTrue=True
        ))

    return sheet_name


# ═══════════════════════════════════════════════════════
# HOJA 3: Observaciones
# ═══════════════════════════════════════════════════════
def _build_observaciones(wb, logs):
    ws = wb.create_sheet("Observaciones")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "HISTORIAL DE OBSERVACIONES"
    c.fill      = _fill(NAVY)
    c.font      = _font(bold=True, color=WHITE, size=13)
    c.alignment = _center()
    ws.row_dimensions[1].height = 30

    headers = [("FECHA", 18), ("ACTIVIDAD", 42), ("ESTADO ANTERIOR", 18),
               ("ESTADO NUEVO", 18), ("USUARIO", 28), ("OBSERVACIÓN", 50)]
    ws.row_dimensions[2].height = 18
    for ci, (h, w) in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill      = _fill(TEAL)
        c.font      = _font(bold=True, color=WHITE, size=10)
        c.alignment = _center(wrap=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    if not logs:
        ws.merge_cells("A3:F3")
        ws["A3"].value     = "No hay observaciones registradas aún."
        ws["A3"].font      = _font(italic=True, color=DKGRAY)
        ws["A3"].alignment = _center()
        return

    for i, log in enumerate(logs):
        r   = 3 + i
        bg  = "FFFFFF" if i % 2 == 0 else "F8FAFC"
        ws.row_dimensions[r].height = 16

        act_name   = "—"
        if log.get("activities"):
            act_name = log["activities"].get("activity_name", "—")
        changed_at = log.get("changed_at", "")[:16].replace("T", " ")
        old_s      = log.get("old_status", "—")
        new_s      = log.get("new_status", "—")
        user       = log.get("user_email", "—")
        obs        = log.get("observation", "") or ""

        row_vals = [changed_at, act_name, old_s, new_s, user, obs]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill      = _fill(bg)
            c.font      = _font(size=9)
            c.alignment = _left(wrap=(ci in [2, 6]))
            c.border    = _border()


# ═══════════════════════════════════════════════════════
# HOJA 4: Instrucciones
# ═══════════════════════════════════════════════════════
def _build_instrucciones(wb):
    ws = wb.create_sheet("Instrucciones")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 70

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value     = "GUÍA DE USO — COMPLIANCE MONITOR SCHEDULE"
    c.fill      = _fill(NAVY)
    c.font      = _font(bold=True, color=WHITE, size=14)
    c.alignment = _center()
    ws.row_dimensions[1].height = 34

    secciones = [
        ("SOBRE ESTE ARCHIVO", [
            "Este archivo es un SNAPSHOT del estado actual del cronograma exportado desde Compliance Monitor.",
            "Puedes seguir trabajando aquí de forma independiente sin necesidad de la app web.",
            "Todos los KPIs del Panel de Control se calculan automáticamente con fórmulas.",
        ]),
        ("CÓMO ACTUALIZAR EL CRONOGRAMA", [
            "1. Ve a la hoja 'Cronograma [Escenario]'.",
            "2. En la columna ESTADO, cada celda tiene un menú desplegable: PENDIENTE → EN PROGRESO → COMPLETADO.",
            "3. Al cambiar el estado, el Gantt de barras se actualiza automáticamente (colores).",
            "4. El Panel de Control se actualiza al presionar F9 (recalcular).",
            "5. Puedes cambiar el RESPONSABLE también con el menú desplegable de esa columna.",
        ]),
        ("SOBRE EL GANTT DE BARRAS", [
            "Las barras de color representan las semanas en que cada actividad debe ejecutarse.",
            "🟩 Verde = COMPLETADO  ·  🟨 Amarillo = EN PROGRESO  ·  🔷 Azul gris = PENDIENTE",
            "Las barras se actualizan automáticamente al cambiar el estado de la columna E.",
        ]),
        ("SOBRE LAS OBSERVACIONES", [
            "La hoja 'Observaciones' muestra el historial completo de cambios registrados en la app.",
            "Es una referencia histórica — puedes añadir tus propias notas en filas nuevas.",
        ]),
        ("SOBRE EL PANEL DE CONTROL", [
            "Los KPIs (completadas, en progreso, pendientes, % avance) son fórmulas COUNTIF.",
            "Apuntan a la columna ESTADO del cronograma — se recalculan solos al editar.",
            "La tabla de avance por fase usa COUNTIFS para filtrar por fase y estado.",
        ]),
        ("COLORES DE REFERENCIA", [
            "Encabezados: Azul Navy (#003A70)  ·  Acentos: Turquesa (#00B5B0)",
            "Completado: Verde (#4ADE80)  ·  En Progreso: Amarillo (#FCD34D)  ·  Pendiente: Azul gris (#CBD5E1)",
        ]),
    ]

    current_row = 3
    for titulo, items in secciones:
        ws.row_dimensions[current_row].height = 10
        current_row += 1

        ws.merge_cells(f"A{current_row}:B{current_row}")
        c = ws.cell(row=current_row, column=1, value=f"  {titulo}")
        c.fill      = _fill(TEAL)
        c.font      = _font(bold=True, color=WHITE, size=11)
        c.alignment = _left(wrap=False)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for item in items:
            ws.cell(row=current_row, column=1, value="")
            c = ws.cell(row=current_row, column=2, value=f"  {item}")
            c.font      = _font(size=10)
            c.alignment = _left(wrap=True)
            ws.row_dimensions[current_row].height = 16
            current_row += 1


# ═══════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ═══════════════════════════════════════════════════════
def generate_excel(activities, logs, kpis, phase_progress, config, scenario):
    wb = Workbook()

    # Nombre de la hoja de cronograma para que las fórmulas del panel la referencien
    sheet_name_crono = f"Cronograma {scenario}"

    _build_panel(wb, kpis, phase_progress, config, scenario, sheet_name_crono)
    _build_cronograma(wb, activities, config, scenario)
    _build_observaciones(wb, logs)
    _build_instrucciones(wb)

    # Eliminar hoja vacía por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
